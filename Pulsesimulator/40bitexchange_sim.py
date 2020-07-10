# -*- coding: utf-8 -*-


import xlrd
import openpyxl as px

fname = '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulsesimulator/pulse_simdata.xlsm'

wb = xlrd.open_workbook(fname)
# print(type(wb))  #Bookオブジェクトを取得
sheet = wb.sheet_by_name('DDSパラメータ')


##########################################################################################
# 入力された周波数（f）を32bitdataに変換

Fout = []
settingF = []

for n in range(1, 10):  # 1だけか、1,10を指定すればDDS1だけかDDS2も使用するか繰り返し回数、つまりRFの個数を指定出来る
    if n % 8 == 1:  # f 1(B),9(J)行目
        # col_valuesの長さは常にsheet.nrowsに等しい 。つまり、sheetの一番下の行数となってしまう
        col_value = sheet.col_values(int(n))
        del col_value[0:3]
        Fout = [s for s in col_value if s != '']  # Fout内の’’を削除
        # print(Fout)
        for m in range(len(Fout)):
            # 設定周波数　＝（2^32/クロック周波数）＊　出力周波数
            setF = (pow(2, 32) / (12.8 * 6)) * int(Fout[m])  # クロック6倍
            setF = int(setF)  # 小数点以下切り捨て
            set32bitF = format(setF, 'b')  # 2進数に変換
            set32bitF = str(set32bitF)
            set32bitF = set32bitF.zfill(32)
            # print(set32bitF)
            settingF.append(set32bitF)
        # print(settingF)

##############################################################################
# 位相制御5bit

Pout = []
settingP = []

for n in range(2, 11):  # 2だけか、2,11を指定すればDDS1だけかDDS2も使用するか繰り返し回数、つまりRFの個数を指定出来る
    if n % 8 == 2:  # phaze 2(C),10(K)行目
        col_value = sheet.col_values(int(n))
        del col_value[0:3]
        Pout = [s for s in col_value if s != '']  # Pout内の’’を削除
        for m in range(len(Pout)):
            θ = int(Pout[m])
            setP = θ / 11.25
            setP = int(setP)
            set5bitP = format(setP, 'b')
            set5bitP = str(set5bitP)
            set5bitP = set5bitP.zfill(5)
            settingP.append(set5bitP)
        # print(settingP)


######################################################################################
# パワーダウン　ロジック　6＊REFの3bit部分

REFout = []
settingREF = []

for n in range(3, 12):  # 3だけか3,12でDDSの繰り返し回数指定
    if n % 8 == 3:
        col_value = sheet.col_values(int(n))
        del col_value[0:3]
        REFout = [s for s in col_value if s != '']
        for l in range(len(REFout)):
            a = int(REFout[l])
            a = str(a)
            settingREF.append(a.zfill(3))
        # print(settingREF)

#####################################################################################

# 各部分の統合
# Phase5bit　REF3bit Frequency32bit　の順番

DDS40bitdata0 = []
DDS40bitdata1 = []  # DDS1の方の40bitdata
DDS40bitdata2 = []  # DDS2の方の40bitdata

for n in range(len(settingP)):
    bitdata40 = settingP[n] + settingREF[n] + settingF[n]
    DDS40bitdata0.append(bitdata40)
# print(DDS40bitdata0)

DDS40bitdata1 = DDS40bitdata0[:len(Fout)]  # スライスで分割
DDS40bitdata2 = DDS40bitdata0[len(Fout): len(settingF)]
# print(DDS40bitdata1)
# print(DDS40bitdata2)

#########################################################################################
# excelファイルに書き込み
# いったんファイルを閉じて開かないと更新されない

wb = px.load_workbook(fname, keep_vba=True)
ws = wb.get_sheet_by_name('DDSパラメータ')


for m in range(len(DDS40bitdata1)):
    ws.cell(column=5, row=(m + 4), value=DDS40bitdata1[m])

for n in range(len(DDS40bitdata2)):
    ws.cell(column=13, row=(n + 4), value=DDS40bitdata2[n])

wb.save(fname)

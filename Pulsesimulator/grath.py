import matplotlib as mtl
import matplotlib.pyplot as plt
import openpyxl as px

fname = '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulsesimulator/pulse_simdata.xlsm'

wb = px.load_workbook(fname, keep_vba=True, data_only=True)
ws = wb.get_sheet_by_name('Pulseパラメータ')


sc1 = []
ec1 = []
i = 1
while ws.cell(row=i, column=1).value != None:
    sc1.append(ws.cell(row=i, column=1).value)
    ec1.append(ws.cell(row=i, column=3).value)
    i += 1
del sc1[0:3]
del ec1[0:3]
# print(sc1)
# print(ec1)

x_pri = sc1 + ec1
x_list = sorted(x_pri)
x_value = sorted(x_list + x_list)
print(x_value)

y_list = [0, 1, 1, 0]
y_value = []
for i in range(len(sc1)):
    y_value += y_list
print(y_value)

plt.plot(x_value, y_value)
plt.title("PL1para")
plt.xlabel("counter")
plt.ylabel('0 or 1')

plt.savefig("hoge.png")  # この行を追記

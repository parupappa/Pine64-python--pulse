# -*- coding: utf-8 -*-


import matplotlib as mtl
import matplotlib.pyplot as plt
import openpyxl as px
import csv
import platform
pf = platform.system()


fname = '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulsesimulator/pulse_simdata.xlsm'

wb = px.load_workbook(fname, keep_vba=True, data_only=True)
ws = wb.get_sheet_by_name('Pulseパラメータ')

#################################################
PLsc1 = []
PLec1 = []
i = 1
while ws.cell(row=i, column=1).value != None:
    PLsc1.append(ws.cell(row=i, column=1).value)
    PLec1.append(ws.cell(row=i, column=3).value)
    i += 1
del PLsc1[0:3]
del PLec1[0:3]
# print(PLsc1)
# print(PLec1)

x_pri = PLsc1 + PLec1
x_list = sorted(x_pri)
x_value1 = sorted(x_list + x_list)

y_list = [0, 1, 1, 0]
y_value1 = []
for i in range(len(PLsc1)):
    y_value1 += y_list


if x_value1[0] != 0:
    x_value1.insert(0, 0)
    y_value1.insert(0, 0)
else:
    pass
# print(x_value1)
# print(y_value1)


##################################################

ADsc1 = []
ADec1 = []
i = 1
while ws.cell(row=i, column=13).value != None:
    ADsc1.append(ws.cell(row=i, column=13).value)
    ADec1.append(ws.cell(row=i, column=15).value)
    i += 1
del ADsc1[0:3]
del ADec1[0:3]
# print(ADsc1)
# print(ADec1)

x_pri = ADsc1 + ADec1
x_list = sorted(x_pri)
x_value4 = sorted(x_list + x_list)

y_value4 = []
for i in range(len(ADsc1)):
    y_value4 += y_list

if x_value4[0] != 0:
    x_value4.insert(0, 0)
    y_value4.insert(0, 0)
else:
    pass
# print(x_value4)
# print(y_value4)


####################################################

dds_csv = open(
    '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulsesimulator/pulse_simDDSdata.csv')

DDS_sc1 = []
DDSsc1_int = []
DDS_data = []
DDS_list = []


for row in csv.reader(dds_csv):
    DDS_sc1.append(row[0])
    DDS_data.append(row[4])
del DDS_sc1[0:3]
del DDS_data[0:3]

# 読み込んだデータをint型に変換
for i in range(len(DDS_sc1)):
    DDSsc1_int.append(int(DDS_sc1[i]))

    # 40bitdataを一文字ずつにして配列に格納
    chars = list(DDS_data[i].strip())
    DDS_list.append(chars)
# 読み込んだ配列をint型に変換
for j in range(len(DDS_list)):
    for i in range(len(DDS_list[0])):
        if str == type(DDS_list[j][i]):
            DDS_list[j][i] = (int(DDS_list[j][i]))
print(DDS_list)
print(DDSsc1_int)


####################################################


x1, y1 = x_value1, y_value1
x4, y4 = x_value4, y_value4
fig = plt.figure()


ax1 = fig.add_subplot(211, title='PL1', yticks=[0, 1], yticklabels=["0", "1"])
ax4 = fig.add_subplot(212, title='AD1',  xlabel='counter', yticks=[
                      0, 1], yticklabels=["0", "1"], sharex=ax1)


ax1.plot(x1, y1)
ax4.plot(x4, y4)

fig.savefig('simulation', facecolor=fig.get_facecolor())

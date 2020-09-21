# -*- coding: utf-8 -*-

# プログラム実装時変更するところ
# for の回数　3ヶ所
# for 例文　range(2) 0 1
#   range(10 ,13 ) 10 11 12


import openpyxl as px

fname = '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulseprogramer/ver2/pulsedata.xlsm'

wb = px.load_workbook(fname)
# print(type(wb))  # Bookオブジェクトを取得
sheet = wb['DDSパラメータ']

##########################################################################################
# 入力された周波数（f）を32bitdataに変換

col_values0 = []
Fout = []
settingF = []
Fout_new = []
f_idx = 0

for n in range(12):
    if n % 6 == 0:
        for cell_obj in list(sheet.columns)[n]:
            col_values0.append(cell_obj.value)
        del col_values0[0:3]
        Fout = [s for s in col_values0 if s != '']  # Fout内の’’を削除
        Fout_new.append(Fout)
        col_values0 = []

while f_idx < 2:
    for frq in Fout_new[f_idx]:
        # 設定周波数 ＝（2^32/クロック周波数）＊ 出力周波数
        setF = (pow(2, 32) // (30 * 6)) * frq  # クロック6倍
        setF = int(setF)  # 小数点以下切り捨て
        set32bitF = format(setF, 'b')  # 2進数に変換
        set32bitF = str(set32bitF).zfill(32)
        # print(set32bitF)
        settingF.append(set32bitF)
        setF, set32bitF = 0, ' '
    f_idx += 1

frq_data1, frq_data2 = settingF[0:8], settingF[8: 16]
frq_data = [frq_data1, frq_data2]
#print(frq_data)

##############################################################################
# 位相制御5bit

Pout = []
Pout_new = []
settingP = []
col_values1 = []
p_idx = 0

for n in range(12):
    if n % 6 == 1:
        for cell_obj in list(sheet.columns)[n]:
            col_values1.append(cell_obj.value)
        del col_values1[0:3]
        Pout = [s for s in col_values1 if s != '']
        Pout_new.append(Pout)
        col_values1 = []

while p_idx < 2:
    for phase in Pout_new[p_idx]:
        θ = phase
        setP = θ / 11.25
        setP = int(setP)
        set5bitP = format(setP, 'b')
        set5bitP = str(set5bitP)
        set5bitP = set5bitP.zfill(5)
        settingP.append(set5bitP)
        setP, set5bitP = 0, ' '
    p_idx += 1

pha_data1, pha_data2 = settingP[0:8], settingP[8:16]
pha_data = [pha_data1, pha_data2]
#print(pha_data)

######################################################################################
# パワーダウン　ロジック　6＊REFの3bit部分

REFout = []
REFout_new = []
settingREF = []
col_values2 = []
REF_idx = 0

for n in range(12):
    if n % 6 == 2:
        for cell_obj in list(sheet.columns)[n]:
            col_values2.append(cell_obj.value)
        del col_values2[0:3]
        REFout = [s for s in col_values2 if s != '']
        REFout_new.append(REFout)
        col_values2 = []

while REF_idx < 2:
    for ref in REFout_new[REF_idx]:
        a = str(ref)
        settingREF.append(a.zfill(3))
    REF_idx += 1

power_data1, power_data2 = settingREF[0:8], settingREF[8:16]
power_data = [power_data1, power_data2]
#print(power_data)
#####################################################################################

# 各部分の統合
# Phase5bit　REF3bit Frequency32bit　の順番

DDS40bitdata0 = []
DDS40bitdata1 = []  # DDS1の方の40bitdata
DDS40bitdata2 = []  # DDS2の方の40bitdata
index = 0

while index < 2:
    for n in range(len(frq_data[index])):
        bitdata40 = pha_data[index][n] + \
            power_data[index][n] + frq_data[index][n]
        DDS40bitdata0.append(bitdata40)
    index += 1
    # print(DDS40bitdata0)


DDS40bitdata1 = DDS40bitdata0[0:8]  # スライスで分割
DDS40bitdata2 = DDS40bitdata0[8: 16]

print(DDS40bitdata1[0])
print(DDS40bitdata1[7])
print(DDS40bitdata2[0])
print(DDS40bitdata2[7])


#########################################################################################
# excelファイルに書き込み
# いったんファイルを閉じて開かないと更新されない

wb = px.load_workbook(fname, keep_vba=True)
sheet = wb['DDSパラメータ']

for m in range(len(DDS40bitdata1)):
    sheet.cell(column=4, row=(m + 4), value=DDS40bitdata1[m])

for n in range(len(DDS40bitdata2)):
    sheet.cell(column=10, row=(n + 4), value=DDS40bitdata2[n])

wb.save(fname)

